import io
import hashlib
import os
import re
import subprocess
import tempfile


OCR_PDF_PAGE_LIMIT = 50
OCR_PAGE_TIMEOUT = 30


def strip_html(text):
    text = re.sub(r"<style[^>]*>.*?</style>", "", text, flags=re.DOTALL)
    text = re.sub(r"<script[^>]*>.*?</script>", "", text, flags=re.DOTALL)
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _ocr_language():
    """Use Bengali too when its Tesseract trained data is installed."""
    configured = os.environ.get("TESSERACT_LANG")
    if configured:
        return configured
    tessdata_dirs = [
        os.environ.get("TESSDATA_PREFIX", ""),
        "/usr/share/tesseract-ocr/5/tessdata",
        "/usr/share/tesseract-ocr/4.00/tessdata",
    ]
    if any(os.path.isfile(os.path.join(directory, "ben.traineddata"))
           for directory in tessdata_dirs if directory):
        return "eng+ben"
    return "eng"


def _ocr_pdf(raw):
    """OCR pages from an image-only PDF when Tesseract is available."""
    try:
        import fitz
    except ImportError:
        return ""

    try:
        doc = fitz.open(stream=raw, filetype="pdf")
    except Exception:
        return ""

    pages = []
    language = _ocr_language()
    print(f"[read_file] Starting PDF OCR with Tesseract language '{language}'")
    try:
        for index, page in enumerate(doc):
            if index >= OCR_PDF_PAGE_LIMIT:
                pages.append("[OCR stopped at the page limit]")
                break
            image_path = None
            try:
                # 2x rendering gives Tesseract enough resolution for typical
                # scanned documents without requiring a separate image library.
                pixmap = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
                with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as image:
                    image.write(pixmap.tobytes("png"))
                    image_path = image.name
                result = subprocess.run(
                    [
                        "tesseract",
                        image_path,
                        "stdout",
                        "--psm",
                        "3",
                        "-l",
                        language,
                    ],
                    capture_output=True,
                    text=True,
                    timeout=OCR_PAGE_TIMEOUT,
                )
                if result.returncode == 0 and result.stdout.strip():
                    pages.append(f"[Page {index + 1}]\n{result.stdout.strip()}")
            except (OSError, subprocess.SubprocessError) as error:
                # Tesseract is optional. Keep any pages already processed so a
                # single failed page does not discard useful OCR output.
                print(f"[read_file] PDF OCR failed on page {index + 1}: {error}")
                break
            finally:
                if image_path:
                    try:
                        os.unlink(image_path)
                    except OSError:
                        pass
    finally:
        doc.close()
    return "\n\n".join(pages)


def _pdf_needs_ocr(page_texts, image_pages):
    """Detect PDFs containing only watermarks or negligible text overlays."""
    nonempty = [text.strip() for text in page_texts if text.strip()]
    if not nonempty or not image_pages:
        return not nonempty and bool(image_pages)
    normalized = {re.sub(r"\W+", " ", text.lower()).strip() for text in nonempty}
    return len(normalized) == 1 or sum(map(len, nonempty)) < len(page_texts) * 40


def _render_pdf_pages(file_path, raw):
    """Render pages as model-readable images when local OCR is unavailable."""
    try:
        import fitz
        doc = fitz.open(stream=raw, filetype="pdf")
    except Exception:
        return []

    digest = hashlib.sha256(raw).hexdigest()[:12]
    directory = os.path.dirname(file_path)
    urls = []
    try:
        for index, page in enumerate(doc):
            if index >= 8:
                break
            filename = f".pdf-page-{digest}-{index + 1}.png"
            try:
                pixmap = page.get_pixmap(matrix=fitz.Matrix(1.5, 1.5), alpha=False)
                pixmap.save(os.path.join(directory, filename))
                urls.append(f"/uploads/{filename}")
            except Exception as error:
                print(f"[read_file] PDF page render failed on page {index + 1}: {error}")
                break
    finally:
        doc.close()
    return urls


def _save_pdf_markdown(file_path, text):
    """Save extracted PDF text as a simple, page-structured Markdown file."""
    markdown_path = os.path.splitext(file_path)[0] + ".ocr.md"
    source_name = os.path.basename(file_path)
    sections = [f"# PDF Extraction: {source_name}", ""]
    for block in text.strip().split("\n\n"):
        match = re.match(r"^\[Page (\d+)\]\n?(.*)$", block, flags=re.DOTALL)
        if match:
            sections.extend([f"## Page {match.group(1)}", "", match.group(2).strip(), ""])
        elif block.strip():
            sections.extend([block.strip(), ""])
    try:
        with open(markdown_path, "w", encoding="utf-8") as output:
            output.write("\n".join(sections).rstrip() + "\n")
        return "/uploads/" + os.path.basename(markdown_path)
    except OSError as error:
        print(f"[read_file] Could not save Markdown output: {error}")
        return None


def _format_ocr_markdown(text):
    """Turn OCR page markers and loose spacing into readable Markdown."""
    blocks = re.split(r"\n\s*\n", text.strip())
    formatted = []
    for block in blocks:
        block = re.sub(r"[ \t]+\n", "\n", block.strip())
        match = re.match(r"^\[Page (\d+)\]\s*(.*)$", block, flags=re.DOTALL)
        if match:
            formatted.append(f"## Page {match.group(1)}")
            if match.group(2).strip():
                formatted.append(match.group(2).strip())
        elif block:
            formatted.append(block)
    return "\n\n".join(formatted)
def read_file_text(file_path):
    ext = os.path.splitext(file_path)[1].lower()
    with open(file_path, "rb") as f:
        raw = f.read()
    if ext == ".pdf":
        try:
            import fitz
            doc = fitz.open(stream=raw, filetype="pdf")
            lines = []
            image_pages = False
            for page in doc:
                lines.append(page.get_text())
                image_pages = image_pages or bool(page.get_images(full=True))
            doc.close()
            text = "\n".join(lines)
            text = strip_html(text)
            if text.strip() and not _pdf_needs_ocr(lines, image_pages):
                markdown_url = _save_pdf_markdown(file_path, text)
                if markdown_url:
                    text += f"\n\n[Markdown saved: {markdown_url}]"
                return text
            if text.strip():
                print("[read_file] PDF text appears to be a watermark/overlay; starting OCR")
            text = _ocr_pdf(raw)
            if text.strip():
                text = _format_ocr_markdown(text)
                markdown_url = _save_pdf_markdown(file_path, text)
                if markdown_url:
                    text += f"\n\n[Markdown saved: {markdown_url}]"
                return text
            page_urls = _render_pdf_pages(file_path, raw)
            if page_urls:
                return (
                    "OCR failed for this image-only PDF. The local OCR engine is "
                    "unavailable or could not recognize any text. You may use "
                    "read_image on the rendered pages as a visual fallback:\n\n"
                    + "\n".join(f"[IMAGE: {url}]" for url in page_urls)
                )
            return (
                "OCR failed for this image-only PDF. The PDF could not be read "
                "because no OCR text was produced."
            )
        except ImportError:
            with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmpf:
                tmpf.write(raw)
                tmp = tmpf.name
            try:
                r = subprocess.run(
                    ["pdftotext", tmp, "-"], capture_output=True, text=True, timeout=30
                )
                return r.stdout
            finally:
                os.unlink(tmp)
    elif ext == ".docx":
        from docx import Document
        doc = Document(io.BytesIO(raw))
        return "\n".join(p.text for p in doc.paragraphs)
    elif ext == ".doc":
        with tempfile.NamedTemporaryFile(suffix=".doc", delete=False) as tmpf:
            tmpf.write(raw)
            tmp = tmpf.name
        try:
            r = subprocess.run(
                ["catdoc", tmp], capture_output=True, text=True, timeout=30
            )
            if r.returncode == 0:
                return r.stdout
            r = subprocess.run(
                ["antiword", tmp], capture_output=True, text=True, timeout=30
            )
            return r.stdout
        finally:
            os.unlink(tmp)
    elif ext in (".xls", ".xlsx"):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        rows = []
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                rows.append("\t".join(str(c) if c is not None else "" for c in row))
        wb.close()
        return "\n".join(rows)
    else:
        # Plain text / code files (.py, .js, .json, .md, .txt, .csv, etc.)
        try:
            print("Reading code file(s)", raw)
            return raw.decode("utf-8")
        except UnicodeDecodeError:
            try:
                print("Trying Latin")
                return raw.decode("latin-1")
            except Exception:
                print("Exception")
                return ""
    return ""
