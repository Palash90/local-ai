"""LLM tool implementations: web search, page fetching, image tools dispatch."""

import concurrent.futures
import json
import os
import re
import threading
import time
from datetime import datetime
from urllib.parse import urlencode, urlparse, urlunparse

import requests
import asyncio

from server.mcp_client import mcp_manager, dispatch_mcp_tool 
from server.features.state import M
from server.features.urlclassify import scrub_search_results
from server.features import page_cache

# How many search results to hand back to the LLM, and how many of the top
# ones to enrich with the full page text (via fetch_page) so the LLM sees real
# content instead of only engine snippets.
WEB_SEARCH_RESULT_LIMIT = 10
WEB_SEARCH_ENRICH_TOP = 6
WEB_SEARCH_ENRICH_CHARS = 6000
WEB_SEARCH_ENRICH_TIMEOUT = 25

# Content-Types treated as plain readable text by fetch_page. Everything else
# (other binary/media types) is declined without ever being fed to the LLM.
_TEXTISH_TYPES = (
    "text/html",
    "text/plain",
    "application/xhtml",
    "application/json",
    "application/xml",
)

# Hard cap on how much text a parsed CSV/spreadsheet/PDF may yield before
# fetch_page stops reading rows/pages, so gigantic documents don't stall or
# blow up the response.
PARSE_ROW_LIMIT = 20000
PARSE_PDF_CHARS = 400000

# When a PDF has no extractable text (scanned pages), render up to this many
# pages to PNG files so the multimodal model can read the page images.
PDF_PAGE_IMAGE_LIMIT = 8
PDF_PAGE_IMAGE_ZOOM = 1.5


def _decode_response_text(raw):
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("latin-1", errors="replace")


def _detect_doc_type(url, content_type, raw):
    """Classify fetched content as ``pdf``/``csv``/``excel`` or ``None``.

    Content-Type is trusted first, then the URL extension, then magic bytes.
    Anything unrecognized returns ``None`` so :func:`fetch_page` declines it as
    binary instead of trying to read it. Legacy ``.xls`` (OLE2 compound
    documents) is detected separately since no parser is available for it.
    """
    ext = os.path.splitext(urlparse(url).path)[1].lower()
    ctype = (content_type or "").split(";")[0].strip().lower()

    if ctype == "application/pdf" or ext == ".pdf" or raw.startswith(b"%PDF-"):
        return "pdf"
    if ctype == "text/csv" or ext == ".csv":
        return "csv"
    # Legacy binary OLE2 compound documents are .xls; we cannot parse those
    # without xlrd, so report them explicitly rather than guessing as binary.
    # Checked before the generic Excel Types because .xls is served as
    # application/vnd.ms-excel just like xlsx templates.
    if ext == ".xls" or raw.startswith(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"):
        return "excel_xls_unsupported"
    if ctype in (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
        "application/excel",
    ) or ext in (".xlsx", ".xlsm", ".xltm", ".xltx"):
        return "excel"
    # Documents served as a generic octet-stream still get parsed if the URL
    # names a known document extension.
    if ctype == "application/octet-stream" and ext in (
        ".pdf",
        ".csv",
        ".xlsx",
        ".xlsm",
        ".xltm",
        ".xltx",
    ):
        return "pdf" if ext == ".pdf" else ("csv" if ext == ".csv" else "excel")
    return None


def _chunk_pages(text, size):
    """Split ``text`` into chunks of up to ``size`` characters."""
    if len(text) <= size:
        return [text]
    return [text[i : i + size] for i in range(0, len(text), size)]


def _doc_result(final_url, title, text, max_chars, chunk=1, page_images=None):
    """Build the ``fetch_page`` JSON payload.

    Long text is split into chunks of up to ``max_chars`` chars and only chunk
    number ``chunk`` is returned; the payload exposes ``total_chunks`` and
    ``next_chunk`` so the agent can page through the rest by calling
    ``fetch_page`` again with ``chunk=2,3,...``. ``page_images`` (rendered PDF
    pages) are attached alongside so a multimodal model can read scanned pages.
    """
    text = (text or "").strip() or "(No readable text content extracted)"
    pages = _chunk_pages(text, max_chars)
    idx = max(0, min(chunk - 1, len(pages) - 1))
    body = pages[idx]
    if page_images:
        body += (
            "\n\n[This PDF has no extractable text — the pages below were rendered "
            "as images. Use the read_image tool on each URL to view a page.]"
        )
    payload = {"url": final_url, "title": title, "content": body}
    if page_images:
        payload["page_images"] = page_images
    if len(pages) > 1:
        payload["chunk"] = idx + 1
        payload["total_chunks"] = len(pages)
        payload["next_chunk"] = idx + 2 if idx + 1 < len(pages) else None
        payload["note"] = (
            f"Page content is split across {len(pages)} chunks. Call fetch_page "
            f"again with chunk={idx + 2} to read the next chunk."
            if payload["next_chunk"]
            else "End of page content."
        )
    return json.dumps(payload, ensure_ascii=False)


def _finish_doc(canon_key, final_url, title, text, max_chars, chunk,
                doc_type="web", page_images=None):
    """Build the fetch payload AND persist the full page under ``canon_key``.

    Only successfully fetched pages reach this point, so the persisted entry is
    guaranteed to have passed the SSRF guard. Pages are stored once with their
    full (pre-chunk) text; every later ``fetch_page(chunk=N)`` re-reads that
    single copy instead of hitting the site again.
    """
    try:
        page_cache.page_put(
            canon_key,
            final_url,
            title or "",
            text or "",
            doc_type=doc_type,
            page_images=page_images,
        )
    except Exception as e:
        print(f"[fetch_page] cache write failed: {e}")
    return _doc_result(final_url, title, text, max_chars, chunk, page_images)


def _render_pdf_pages(doc, url):
    """Render image-only PDF pages (scanned) to PNG files the model can view.

    Rendered files land under ``IMG_PATH/pdf_pages`` and are returned as
    ``/output/pdf_pages/...`` URLs so ``read_image`` / ``resolve_image_path``
    can resolve them. Returns an empty list when nothing could be rendered.
    """
    import hashlib

    slug = hashlib.md5(url.encode("utf-8", errors="replace")).hexdigest()[:12]
    outdir = os.path.join(M.IMG_PATH, "pdf_pages")
    try:
        os.makedirs(outdir, exist_ok=True)
    except OSError as e:
        print(f"[fetch_page] PDF page render dir failed: {e}")
        return []
    urls = []
    for i, page in enumerate(doc):
        if i >= PDF_PAGE_IMAGE_LIMIT:
            break
        try:
            import fitz

            pix = page.get_pixmap(
                matrix=fitz.Matrix(PDF_PAGE_IMAGE_ZOOM, PDF_PAGE_IMAGE_ZOOM)
            )
            fname = f"{slug}-p{i + 1}.png"
            pix.save(os.path.join(outdir, fname))
            urls.append(f"/output/pdf_pages/{fname}")
        except Exception as e:
            print(f"[fetch_page] PDF page {i + 1} render failed: {e}")
            break
    return urls


def _parse_pdf(raw, url):
    """Extract text from a PDF via PyMuPDF.

    Returns ``(text, title, page_images)``. If the PDF has no extractable text
    (scanned pages) the first ``PDF_PAGE_IMAGE_LIMIT`` pages are rendered to PNG
    files and returned as ``/output/pdf_pages/...`` URLs.
    """
    doc_title = os.path.basename(urlparse(url).path) or "PDF document"
    try:
        import fitz

        doc = fitz.open(stream=raw, filetype="pdf")
        parts = []
        total_chars = 0
        for page in doc:
            text = page.get_text("text")
            parts.append(text)
            total_chars += len(text)
            if total_chars > PARSE_PDF_CHARS:
                parts.append("\n...[truncated by size]")
                break
        body = "\n".join(parts).strip()
        page_images = []
        if not body:
            page_images = _render_pdf_pages(doc, url)
            body = "(No extractable text in PDF — the pages are likely scanned images.)"
        doc.close()
        return body, doc_title, page_images
    except Exception as e:
        print(f"[fetch_page] PDF parse failed: {e}")
        return f"(Could not extract text from this PDF: {e})", doc_title, []


def _parse_csv(raw, url):
    """Parse CSV text into pipe-separated rows. Returns ``(text, title)``."""
    import csv
    import io

    doc_title = os.path.basename(urlparse(url).path) or "CSV document"
    try:
        reader = csv.reader(io.StringIO(_decode_response_text(raw)))
        rows = []
        for i, row in enumerate(reader):
            if i >= PARSE_ROW_LIMIT:
                rows.append("...[truncated by row limit]")
                break
            rows.append(" | ".join("" if c is None else c.strip() for c in row))
        body = "\n".join(rows).strip() or "(Empty CSV)"
        return body, doc_title
    except Exception as e:
        print(f"[fetch_page] CSV parse failed: {e}")
        return f"(Could not parse this CSV: {e})", doc_title


def _parse_excel(raw, url):
    """Extract every sheet of an .xlsx workbook as text rows. Returns (text, title)."""
    import io

    from openpyxl import load_workbook

    doc_title = os.path.basename(urlparse(url).path) or "Excel spreadsheet"
    try:
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        blocks = []
        for sheet in wb.worksheets:
            blocks.append(f"### Sheet: {sheet.title}")
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                if row_idx >= PARSE_ROW_LIMIT:
                    blocks.append("...[truncated by row limit]")
                    break
                blocks.append(" | ".join("" if v is None else str(v) for v in row))
        wb.close()
        body = "\n".join(blocks).strip() or "(Empty spreadsheet)"
        return body, doc_title
    except Exception as e:
        print(f"[fetch_page] Excel parse failed: {e}")
        return f"(Could not parse this spreadsheet: {e})", doc_title


# Search-result cache and outbound pacing for web_search. Non-time-sensitive
# results are reused for a day for exact and near-duplicate queries;
# time-sensitive queries are only reused within a short window. Outbound
# fetches are paced globally so bursts from the research agents cannot get
# the upstream engines rate-limited or CAPTCHA'd. Cache reads take only
# _CACHE_LOCK and never wait on pacing or on an in-flight fetch.
SEARCH_CACHE_MAX = 256
SEARCH_SIMILARITY_THRESHOLD = 0.7
SEARCH_MIN_INTERVAL = 5.0
SEARCH_INFLIGHT_WAIT = 30
_CACHE_LOCK = threading.Lock()
_PACE_LOCK = threading.Lock()
_SEARCH_LAST_FETCH = 0.0
_SEARCH_CACHE = {}
_IN_FLIGHT = {}


def _search_cache_get(norm_query, query, now):
    """Lock-free cache lookup; caller must hold _CACHE_LOCK."""
    ttl = page_cache.regex_ttl(query)
    entry = _SEARCH_CACHE.get(norm_query)
    if entry:
        # Prefer the per-entry TTL fixed when the answer was cached (which may
        # have come from the LLM classifier); fall back to the regex default.
        age = now - entry[0]
        if age <= (entry[2] if len(entry) > 2 else ttl):
            payload = dict(entry[1])
            payload["cached_result"] = True
            return payload
    tokens = set(norm_query.split())
    if not tokens:
        return None
    best_key, best_score = None, 0.0
    for key, (cached_at, _, store_ttl) in _SEARCH_CACHE.items():
        if now - cached_at > store_ttl:
            continue
        cand = set(key.split())
        if not cand:
            continue
        score = len(tokens & cand) / min(len(tokens), len(cand))
        if score > best_score:
            best_key, best_score = key, score
    if best_score >= SEARCH_SIMILARITY_THRESHOLD:
        payload = dict(_SEARCH_CACHE[best_key][1])
        payload["cached_result"] = True
        return payload
    return None


def _search_cache_store(norm_query, payload, ttl=None):
    snapshot = json.loads(json.dumps(payload))
    if ttl is None:
        ttl = page_cache.regex_ttl(norm_query)
    with _CACHE_LOCK:
        _SEARCH_CACHE[norm_query] = (time.monotonic(), snapshot, ttl)
        while len(_SEARCH_CACHE) > SEARCH_CACHE_MAX:
            oldest = min(_SEARCH_CACHE, key=lambda k: _SEARCH_CACHE[k][0])
            del _SEARCH_CACHE[oldest]


def _finish_inflight(norm_query, event):
    event.set()
    with _CACHE_LOCK:
        if _IN_FLIGHT.get(norm_query) is event:
            del _IN_FLIGHT[norm_query]


def _pace_outbound_request():
    """Space outbound fetches at least SEARCH_MIN_INTERVAL apart.

    Holds _PACE_LOCK while sleeping so concurrent fetchers queue up behind
    the pacing decision; cache readers are never blocked by this lock.
    """
    global _SEARCH_LAST_FETCH
    with _PACE_LOCK:
        wait = SEARCH_MIN_INTERVAL - (time.monotonic() - _SEARCH_LAST_FETCH)
        if wait > 0:
            time.sleep(wait)
        _SEARCH_LAST_FETCH = time.monotonic()


# Semantic web_search reuse. When a query has no exact keyed-cache match we ask
# the vector layer for previously-fetched pages whose meaning resembles the
# query, then reuse them as the result set. Only high-confidence hits are used;
# below this score the query degrades to a live SearXNG fetch.
_SEMANTIC_HIT_MIN_SCORE = 0.60
_SEMANTIC_HIT_K = 5
# Semantic page recall is unsafe for live or location-sensitive requests: a
# cached London traffic page is semantically close to Kolkata traffic while
# being factually useless.
_SEMANTIC_RECALL_UNSAFE_RE = re.compile(
    r"\b(?:traffic|weather|nearby|local|live|current(?:ly)?|now|latest|"
    r"breaking|today|tonight|news|updates?)\b",
    re.IGNORECASE,
)

# TTL the LLM is allowed to return, in seconds.
_LLM_TTL_MIN = 60
_LLM_TTL_MAX = 30 * 24 * 3600
# Few-shot prompt for the raw /completion endpoint. The chat models here think
# unconditionally (channel templates) and burn the token budget before the
# JSON ever appears, so the classifier uses base-style completion instead:
# examples teach the judgment, a GBNF grammar guarantees the shape, and the
# shared prefix stays prompt-cached (cache_reuse) for ~0.5s classifications.
_LLM_TTL_FEWSHOT = (
    "Classify how long a web-search answer stays valid.\n"
    'Query: latest breaking news headlines -> {"ttl_seconds": 300}\n'
    'Query: how to bake sourdough bread -> {"ttl_seconds": 2592000}\n'
    'Query: nvidia stock price right now -> {"ttl_seconds": 60}\n'
    'Query: python asyncio tutorial -> {"ttl_seconds": 2592000}\n'
    'Query: premier league results from yesterday -> {"ttl_seconds": 3600}\n'
)
_LLM_TTL_GRAMMAR = (
    'root ::= "{" [ ]*"\\\"ttl_seconds\\\""[ ]*":"[ ]*[0-9]+ "}"'
)
_LLM_TTL_TIMEOUT = 10
# Lanes tried in order for the TTL classification. The GPU lane serves every
# interactive chat (web_search's only caller), the CPU/guardrail lanes serve
# agent runs. All three sit behind lazy routers that would happily START a
# multi-GB model load for our little request, so each lane is gated on
# ``is_model_ready`` first: an unloaded lane is skipped, never poked.
_LLM_TTL_LANES = ("gpu", "guardrail", "cpu")


def _llm_ttl(query, default_ttl):
    """Ask the LLM how long this query's answer stays fresh.

    Returns seconds for the cache. On any failure (no loaded lane, malformed
    reply) falls back to ``default_ttl`` so a classifier hiccup never blocks a
    search. Uses grammar-constrained raw completion: one tiny decode pass,
    immune to thinking-mode token burn. Only lanes with an already-loaded
    model are used — this call must never itself trigger a model load.
    """
    payload = {
        "prompt": _LLM_TTL_FEWSHOT + f"Query: {query} ->",
        "n_predict": 16,
        "temperature": 0.1,
        "grammar": _LLM_TTL_GRAMMAR,
        "cache_reuse": 256,
    }
    r = None
    for lane in _LLM_TTL_LANES:
        url = M.server_url(lane)
        base = url[: -len("/v1/chat/completions")] if url.endswith(
            "/v1/chat/completions"
        ) else url
        model_id = M.server_model_id(lane)
        if not M.is_model_ready(base, model_id):
            continue
        try:
            r = requests.post(
                f"{base}/completion",
                json=dict(payload, model=model_id),
                timeout=_LLM_TTL_TIMEOUT,
            )
            if r.status_code == 200:
                break
            print(f"[ttl] {lane} lane HTTP {r.status_code}: {r.text[:120]}")
        except Exception as e:
            print(f"[ttl] {lane} lane unavailable: {e}")
    else:
        print("[ttl] no loaded LLM lane, using regex default")
        return default_ttl
    try:
        content = r.json().get("content") or ""
        start, end = content.find("{"), content.rfind("}")
        if start == -1 or end == -1:
            print(f"[ttl] classifier non-JSON: {content[:80]!r}")
            return default_ttl
        raw = int(json.loads(content[start : end + 1]).get("ttl_seconds") or 0)
        if raw <= 0:
            print(f"[ttl] classifier returned non-positive TTL: {raw!r}")
            return default_ttl
    except Exception as e:
        print(f"[ttl] classifier failed, using default: {e}")
        return default_ttl
    ttl = max(_LLM_TTL_MIN, min(_LLM_TTL_MAX, raw))
    print(f"[ttl] LLM TTL for {query!r} -> {ttl}s")
    return ttl


def _semantic_search_hit(query, min_score=_SEMANTIC_HIT_MIN_SCORE,
                         k=_SEMANTIC_HIT_K):
    """Return a web_search payload built from vector-cached pages, or None.

    ``page_semantic`` returns pages already fetched/stored (with real nomic
    embeddings). If the top hits clear ``min_score`` we can answer the query
    from them instead of the network.
    """
    if _SEMANTIC_RECALL_UNSAFE_RE.search(query or ""):
        return None
    hits = page_cache.page_semantic(query, k=k, min_score=min_score)
    if not hits:
        return None
    results = []
    for h in hits:
        results.append(
            {
                "title": h["title"] or h["url"],
                "url": h["url"],
                "snippet": h.get("snippet") or "",
                "semantic_score": h["score"],
            }
        )
    if not results:
        return None
    payload = {
        "results": results,
        "search_date": datetime.now().strftime("%Y-%m-%d %A"),
        "query": query,
        "semantic_recall": True,
    }
    return payload


_CATEGORY_RE = (
    (re.compile(r"news|updates?\b|today|latest|breaking|announ[ce]|unveil",
                re.I), "news"),
    (re.compile(r"python|javascript|typescript|github|gitlab|git\b|docker|"
                r"kubernetes|linux|unix|apache|nginx|devops|cloud|"
                r"programming|developer|open.?source|bug\b|debug|compile|"
                r"deploy|server|database|sql\b|terminal|command.?line|script|"
                r"django|flask|react|angular|node\b|rails|laravel|spring\b|"
                r"tensorflow|pytorch|\bAI\b|\bLLM\b|model\s+weights|error",
                re.I), "it"),
    (re.compile(r"arxiv|paper|research|study|journal|scientific|physics|"
                r"mathematics?|mathematical|chemistry|chemical|biology|"
                r"genome|quantum|neuron|astronom|cosmolog|doi\b|experiment|"
                r"hypothesis|theorem|calculus|peer.?review", re.I), "science"),
)


def _pick_categories(query):
    """Best-guess SearXNG categories for a query.

    Returns ``general`` plus any specific category the query clearly matches
    (news/it/science), so a search never gets locked to a single narrow engine
    pool — a bare ``categories=it`` only searches github/stackoverflow and can
    return empty results for queries that merely mention an ``AI``-style token.
    """
    cats = ["general"]
    matched = [cat for rx, cat in _CATEGORY_RE if rx.search(query or "")]
    cats += [cat for cat in ("news", "it", "science") if cat in matched]
    return ",".join(cats)


# ---------------------------------------------------------------------------
# Result relevance gating.
#
# SearXNG (especially the local instance) does naive token matching and will
# happily return keyword-only junk for a multi-word query — the "real-time
# traffic" → Real Madrid and "quantum computing overview" → dictionary-of-the-
# -word-`overview` failures seen in production. Two layers run on every fresh
# search before results reach the LLM:
#   1. lexical: drop results sharing no meaningful content token with the query;
#   2. semantic: score each surviving result against the query with the local
#      nomic embedding server and drop low-similarity hits.
# Both degrade gracefully: if the tokenizer/embed server is unavailable we keep
# the results rather than empty the list, but we still surface ``low_confidence``
# so the LLM is told to treat the set as unreliable instead of improvising.
# ---------------------------------------------------------------------------

# Semantic similarity (cosine 0..1) a result must reach to be considered
# credible. Genuinely relevant hits sit 0.5-0.85; the keyword-only junk cases
# (Real Madrid vs traffic scored ~0.37, dictionary page vs quantum ~0.2) sit
# below it, so 0.40 is a deliberate gap that drops them while keeping real hits.
REL_MIN_SEMANTIC = 0.40
# If fewer than this many results survive BOTH gates we surface the (possibly
# empty) survivor set with ``low_confidence=True`` and a re-search directive so
# the model never fixates on an irrelevant single result (e.g. Real Madrid in
# a Bangalore traffic query).
REL_MIN_CREDIBLE_RESULTS = 2

# Content words that deserve no query token (role-filler / vague-intent words
# that match unrelated dictionary pages, e.g. "overview", "research", "report").
_QUERY_STOPWORDS = set(
    """
    a an the of on in to for and or but about with at by from via
    what who when where why how is are was were be been being do does did
    can could will would should shall may might must have has had
    this that these those it its there here
    overview summary general basics fundamental fundamentals introduction
    detail details comprehensive research report find information info explain
    explainer guide explain write about
    """.split()
)


def _query_tokens(query):
    """Lowercased, stopword-stripped meaningful query terms.

    Hyphenated compounds are kept intact (``real-time`` stays ONE token so the
    query never over-matches the word ``real``, e.g. against "Real Madrid" in a
    traffic search); underscores are collapsed to plain tokens.
    """
    q = (query or "").lower()
    q = q.replace("_", " ")
    toks = set(re.findall(r"[a-z0-9]+(?:-[a-z0-9]+)*", q))
    return toks - _QUERY_STOPWORDS


def _result_tokens(entry):
    """Lowercased meaningful tokens from a result's title + snippet."""
    text = " ".join(
        [
            entry.get("title", "") or "",
            entry.get("page_title", "") or "",
            entry.get("snippet", "") or "",
        ]
    ).lower()
    text = text.replace("_", " ")
    return set(re.findall(r"[a-z0-9]+(?:-[a-z0-9]+)*", text)) or set()


# High-frequency, highly ambiguous English words that routinely cause false
# keyword "matches" (e.g. the standalone "real" from a space-separated
# "real time" query colliding with "Real Madrid"). A result whose ONLY overlap
# with the query is on these weak tokens counts as having NO overlap at all.
_WEAK_TOKENS = {
    "real", "top", "best", "new", "latest", "more", "today", "result",
    "results", "big", "info", "information", "page", "main", "item", "items",
    "list", "article",
}


def _strong_overlap(qtoks, rtoks):
    """True if the token overlap includes at least one non-weak query token."""
    overlap = qtoks & rtoks
    if not overlap:
        return False
    strong = qtoks - _WEAK_TOKENS
    return bool(overlap & strong)


_PLACE_ALIASES = {
    "bengaluru": ("bengaluru", "bangalore"),
    "bangalore": ("bengaluru", "bangalore"),
    "kolkata": ("kolkata", "calcutta"),
    "calcutta": ("kolkata", "calcutta"),
    "mumbai": ("mumbai", "bombay"),
    "bombay": ("mumbai", "bombay"),
    "united states": ("united states", "usa", "us"),
    "usa": ("united states", "usa", "us"),
    "us": ("united states", "usa", "us"),
}


def _requested_places(query):
    """Return explicitly named places in a query, including known aliases."""
    q = (query or "").lower()
    places = []
    for place in _PLACE_HINTS:
        if re.search(rf"(^| ){re.escape(place)}( |$)", q):
            places.extend(_PLACE_ALIASES.get(place, (place,)))
    return set(places)


def _location_matches(query, entry):
    """Reject a result that cannot be about an explicitly requested place."""
    requested = _requested_places(query)
    if not requested:
        return True
    text = " ".join(
        str(entry.get(key, "") or "")
        for key in ("title", "page_title", "snippet")
    ).lower()
    url = str(entry.get("url", "") or "").lower()
    return any(
        re.search(rf"(^|[^a-z]){re.escape(place)}([^a-z]|$)", text)
        or place.replace(" ", "") in url
        for place in requested
    )


def _cosine(a, b):
    if not a or not b or len(a) != len(b):
        return 0.0
    dot = sum(x * y for x, y in zip(a, b))
    na = sum(x * x for x in a) ** 0.5
    nb = sum(y * y for y in b) ** 0.5
    if na == 0.0 or nb == 0.0:
        return 0.0
    return dot / (na * nb)


def _semantic_relevance(results, query):
    """Score results against the query with the embed server.

    Returns (results, low_confidence). Results below ``REL_MIN_SEMANTIC`` are
    HARD-DROPPED (never kept just so the list is non-empty — a surviving junk
    result like a Real Madrid page makes the model latch onto it). On embed
    failure returns (results, False) untouched so the caller keeps the lexical
    survivors rather than emptying the list over a transient outage.
    """
    if not results:
        return results, False
    texts = [query] + [
        f"{r.get('title') or ''} :: {r.get('snippet') or ''}"[:280] for r in results
    ]
    vecs = page_cache.embed_texts(texts)
    if not vecs or len(vecs) != len(texts):
        return results, False
    qv = vecs[0]
    scored = []
    for r, rv in zip(results, vecs[1:]):
        r["relevance"] = round(_cosine(qv, rv), 3)
        if r["relevance"] >= REL_MIN_SEMANTIC:
            scored.append(r)
    scored.sort(key=lambda r: r.get("relevance", 0.0), reverse=True)
    low_conf = len(scored) < REL_MIN_CREDIBLE_RESULTS
    return scored[:WEB_SEARCH_RESULT_LIMIT], low_conf


def _filter_relevant_results(results, query):
    """Apply lexical + semantic relevance gating; return (results, low_confidence)."""
    if not results:
        return results, False
    qtoks = _query_tokens(query)
    # Location gate is authoritative for explicit place queries. Generic
    # topical similarity must not turn a Kolkata query into a London/TfL hit.
    location_kept = [r for r in results if _location_matches(query, r)]
    if _requested_places(query):
        if not location_kept:
            return [], True
        results = location_kept
    # Lexical gate is AUTHORITATIVE: a result sharing no meaningful token with
    # the query (Real-Madrid-for-"real-time traffic", dictionary-for-"overview")
    # is junk and is hard-dropped — never handed back just to avoid an empty
    # list, because the model will then fixate on the irrelevant page.
    if qtoks:
        kept = [r for r in results if _strong_overlap(qtoks, _result_tokens(r))]
        if kept:
            results = kept
        else:
            return [], True
    # Semantic gate: drop low-similarity survivors; degrade gracefully to the
    # lexical survivors if the embed server is briefly unavailable.
    try:
        results, low_conf = _semantic_relevance(results, query)
    except Exception as e:
        print(f"[web_search] relevance screening failed: {e}")
        results, low_conf = results, False
    return results[:WEB_SEARCH_RESULT_LIMIT], low_conf


def _screen_cached_payload(payload, query):
    """Revalidate cached results because cache keys may predate relevance gates."""
    if not isinstance(payload, dict) or not isinstance(payload.get("results"), list):
        return payload
    screened = dict(payload)
    results, low_confidence = _filter_relevant_results(payload["results"], query)
    screened["results"] = results
    if low_confidence:
        screened["low_confidence"] = True
        screened["low_confidence_note"] = (
            "Cached search results did not contain enough credible sources for "
            "this query. Do not fetch or cite the listed URLs; issue a fresh, "
            "better-scoped web_search instead."
        )
    return screened


# ---------------------------------------------------------------------------
# Query re-scoping for the search backend and the diagnostics fallback fetch.
# ---------------------------------------------------------------------------
_COUNTRY_LANG = [
    (("united states", "usa", "america"), "en-US"),
    (("us",), "en-US"),
    (("india", "bengaluru", "bangalore", "mumbai", "delhi", "kolkata", "chennai", "hyderabad", "pune"), "en-IN"),
    (("united kingdom", "uk", "britain", "london"), "en-GB"),
    (("australia", "sydney", "melbourne"), "en-AU"),
    (("canada", "toronto"), "en-CA"),
]


def _region_language(query, location=""):
    """Best-guess SearXNG ``language`` code from a query's geo tokens + location."""
    q = " ".join(_query_tokens(query))
    q = q.replace("-", " ")
    if q:
        for keys, lang in _COUNTRY_LANG:
            for k in keys:
                if re.search(rf"(^| ){re.escape(k)}( |$)", q):
                    return lang
    loc = (location or "").lower()
    if "india" in loc or "bengaluru" in loc or "bangalore" in loc:
        return "en-IN"
    return ""


def _rescoped_query(query):
    """Drop role-filler words so the backend/fallback searches real content.

    E.g. "overview of quantum computing fundamentals and applications" →
    "quantum computing", which the local SearXNG can actually answer instead
    of returning dictionary pages for the word "overview".
    """
    toks = _query_tokens(query)
    return " ".join(sorted(toks, key=str.lower)) if toks else (query or "")


def _apply_location_scoping(clean_query, current_location, params):
    """Fold location/geo into SearXNG params (language hint, place qualifier)."""
    lang = _region_language(clean_query, current_location)
    if lang:
        params["language"] = lang
    # If the user gave no explicit place and we know where they are AND the
    # query is location-sensitive, qualify the search with their area so the
    # engine returns local results instead of global keyword-only junk.
    place = (current_location or "").strip()
    if place and not _has_explicit_place(clean_query):
        if re.search(r"weather|traffic|news|nearby|restaurants?|caf[eé]s?|"
                     r"events?|local|today|forecast|store|shop|market", clean_query, re.I):
            params["q"] = f"{clean_query} {place}"


def _has_explicit_place(query):
    """True when the query already names a city/country/place."""
    return any(re.search(rf"(^| ){re.escape(place)}( |$)",
                         query.lower())
               for place in _PLACE_HINTS)


_PLACE_HINTS = [
    "bengaluru", "bangalore", "mumbai", "delhi", "kolkata", "chennai",
    "hyderabad", "pune", "india", "london", "paris", "new york", "tokyo",
    "boston", "seattle", "sydney", "melbourne", "toronto", "us", "usa",
    "united states", "america", "uk", "singapore", "bangkok", "dubai",
]


def web_search(query, current_time=None, current_location=None):
    ts = datetime.now()
    clean_query = (query or "").strip()
    norm_query = " ".join(re.findall(r"[a-z0-9]+", clean_query.lower()))
    # Live and location-sensitive searches must not reuse an older result set,
    # including one that may have been contaminated by semantic recall.
    allow_cached_results = not _SEMANTIC_RECALL_UNSAFE_RE.search(clean_query)
    params = {"q": clean_query, "format": "json"}
    _apply_location_scoping(clean_query, current_location, params)
    cats = _pick_categories(clean_query)
    if cats:
        params["categories"] = cats
    print(f"[web_search] categories={cats!r} for {clean_query!r}")
    # The diagnostics fallback never repeats a query the engine already failed:
    # it re-runs the ROLE-FILLER-STRIPPED content words (plus geo scope) so a
    # "quantum computing overview" dead-end becomes a real "quantum computing"
    # search instead of dictionary pages for the word "overview". It is still
    # explicitly non-citable.
    fallback_params = dict(params)
    fallback_params["q"] = _rescoped_query(clean_query)
    _apply_location_scoping(fallback_params["q"], current_location, fallback_params)
    fallback_url = f"{M.SEARXNG_PUBLIC_URL}?{urlencode(fallback_params)}"

    def _respond(results, error=None, low_confidence=False):
        payload = {
            "results": results,
            "search_date": ts.strftime("%Y-%m-%d %A"),
            "retrieved_at": ts.astimezone().isoformat(timespec="seconds"),
            "query": query,
            "fallback_fetch_url": fallback_url,
            "fallback_fetch_note": (
                "Diagnostics-only re-run of the query with role-filler words "
                "stripped (e.g. 'overview'/'fundamentals' removed) plus the "
                "geo/language scope, returned as raw JSON. NEVER cite this URL "
                "as a source; cite result URLs only. Prefer issuing a fresh, "
                "well-scoped web_search instead of fetching this, and only use "
                "it to confirm whether the engine has better results."
            ),
        }
        if low_confidence:
            payload["low_confidence"] = True
            payload["low_confidence_note"] = (
                "This search FAILED to find any result credibly matching the "
                "query (fewer than two results survived relevance screening; "
                "the set may even be empty). Treat every listed URL as "
                "UNRELIABLE and off-topic: do NOT fetch any of them, do NOT "
                "summarize them, and do NOT build an answer around them (e.g. "
                "do not pivot to Real Madrid for a traffic question just "
                "because it is the only returned link). State the sub-answer "
                "as UNSUPPORTED and issue a fresh web_search with a clearer, "
                "better-scoped query instead of improvising or latching onto "
                "an irrelevant source."
            )
        if error:
            payload["error"] = error
        return payload

    owns_slot = False
    with _CACHE_LOCK:
        hit = (
            _search_cache_get(norm_query, clean_query, time.monotonic())
            if allow_cached_results
            else None
        )
        if hit is None and allow_cached_results:
            inflight = _IN_FLIGHT.get(norm_query)
            if inflight is None:
                inflight = _IN_FLIGHT[norm_query] = threading.Event()
                owns_slot = True
    if hit is not None:
        print("Web-search cache hit")
        return json.dumps(_screen_cached_payload(hit, clean_query))
    if hit is None and allow_cached_results:
        # Persistent (cross-restart) cache: a query answered before this
        # process started should never cost another SearXNG request.
        hit = page_cache.search_get(norm_query)
        if hit is not None:
            _search_cache_store(norm_query, hit, ttl=hit.get("_ttl"))
            print("Web-search cache hit (persistent)")
            return json.dumps(_screen_cached_payload(hit, clean_query))
    if hit is None and allow_cached_results:
        # Semantic recall: this query shares no exact key with a cached search,
        # but we may have already fetched pages (via fetch_page or enrichment)
        # whose meaning matches the query. Reuse those instead of another
        # SearXNG call — the whole point of the vector layer.
        hit = _semantic_search_hit(clean_query)
        if hit is not None:
            if owns_slot:
                _finish_inflight(norm_query, inflight)
            _search_cache_store(norm_query, hit)
            print("Web-search cache hit (semantic)")
            return json.dumps(_screen_cached_payload(hit, clean_query))
    if not owns_slot and allow_cached_results:
        # An identical fetch is already running: wait for its result
        # instead of duplicating the request.
        if inflight.wait(SEARCH_INFLIGHT_WAIT):
            with _CACHE_LOCK:
                hit = _search_cache_get(
                    norm_query, clean_query, time.monotonic()
                )
            if hit is not None:
                print("Web-search cache hit (in-flight wait)")
                return json.dumps(_screen_cached_payload(hit, clean_query))
        # Timed out or the fetch failed without caching: fall through and
        # fetch this query ourselves (paced like any other).

    _pace_outbound_request()
    print("Performing web search", M.SEARXNG_URL)
    try:
        r = requests.get(M.SEARXNG_URL, params=params, timeout=10)
        r.raise_for_status()
        data = r.json()
    except Exception as e:
        print(f"Web-search failed: {e}")
        _finish_inflight(norm_query, inflight)
        return json.dumps(_respond([], error=str(e)))

    results = data.get("results", [])[:WEB_SEARCH_RESULT_LIMIT]
    formatted = []
    for x in results:
        formatted.append(
            {
                "title": x.get("title", ""),
                "url": x.get("url", ""),
                "snippet": x.get("content", "") or x.get("snippet", ""),
            }
        )
    # Scrub ads/promo and orphan landing pages BEFORE this result set is handed
    # to the LLM (it is also what gets recorded as _search_details and later
    # re-read by the story pipeline). The model should never see a shopping,
    # download, quote, or company-profile URL as a possible source.
    formatted = scrub_search_results(formatted, query=query)
    # Relevance gate (A): drop keyword-only junk (Real-Madrid-for-traffic,
    # dictionary-for-"overview") before the LLM ever sees it, and flag the
    # set as low-confidence when few credible results survive so the model
    # reports UNSUPPORTED instead of improvising.
    formatted, low_confidence = _filter_relevant_results(formatted, query)
    payload = _respond(formatted, low_confidence=low_confidence)
    # Ask the LLM how long this answer stays fresh so the next identical query
    # re-fetches at the right time ("breaking news" -> seconds, "how to" -> days),
    # rather than a fixed regex heuristic. Falls back to the regex default if
    # the classifier call fails or the LLM is unreachable. Computed once and
    # used for both the in-memory and persistent stores.
    ttl = _llm_ttl(query, page_cache.regex_ttl(query))
    # Cache the raw results before enrichment so waiters can pick them up,
    # then re-store the enriched version once page fetching completes.
    _search_cache_store(norm_query, payload, ttl=ttl)
    _finish_inflight(norm_query, inflight)
    enriched = _enrich_top_results(formatted)
    payload["results"] = enriched
    _search_cache_store(norm_query, payload, ttl=ttl)
    page_cache.search_put(norm_query, query, payload, ttl=ttl)
    return json.dumps(payload)


def _enrich_top_results(results):
    """Attach the full page text to the top results.

    The top ``WEB_SEARCH_ENRICH_TOP`` results are fetched concurrently (their
    body is stored as ``full_content``) so the LLM does not have to make a
    separate ``fetch_page`` call for every promising link. Fetch failures are
    recorded as ``fetch_error`` and never break the search response.
    """
    targets = results[:WEB_SEARCH_ENRICH_TOP]
    if not targets:
        return results

    def _one(entry):
        url = entry.get("url", "")
        if not url.lower().startswith(("http://", "https://")):
            return
        try:
            page = json.loads(M.fetch_page(url, max_chars=WEB_SEARCH_ENRICH_CHARS))
            if page.get("content"):
                entry["full_content"] = page["content"]
                entry["page_title"] = page.get("title", "")
            elif page.get("error"):
                entry["fetch_error"] = page["error"]
        except Exception as e:
            entry["fetch_error"] = str(e)

    executor = concurrent.futures.ThreadPoolExecutor(max_workers=len(targets))
    futs = [executor.submit(_one, entry) for entry in targets]
    concurrent.futures.wait(futs, timeout=WEB_SEARCH_ENRICH_TIMEOUT)
    executor.shutdown(wait=False)
    return results


def _rewrite_search_url_to_internal(parsed):
    """Map the public SearXNG search URL onto the internal SearXNG instance.

    The public host resolves to a loopback address via /etc/hosts on this
    machine, so fetch_page's SSRF guard would refuse to fetch the very search
    URL that web_search hands out as a citation. Returns None for URLs that
    are not the search endpoint.
    """
    public = urlparse(M.SEARXNG_PUBLIC_URL or "")
    internal = urlparse(M.SEARXNG_URL or "")
    if not (public.netloc and internal.scheme and internal.netloc):
        return None
    base = (public.path or "").rstrip("/")
    path = parsed.path or "/"
    if path != base and not path.startswith(base + "/"):
        return None
    prefix = (internal.path or "").rstrip("/")
    if prefix and path.startswith(prefix):
        path = path[len(prefix):] or "/"
    return urlunparse(internal._replace(path=path, query=parsed.query))


def _republicize_internal_url(payload_json):
    """Replace the internal SearXNG URL with the public one in a tool payload.

    fetch_page rewrites the public search URL onto the internal instance to
    pass the SSRF guard, but the model (and the user) must only ever see the
    public URL — otherwise the private backend address leaks into citations.
    """
    internal = (M.SEARXNG_URL or "").rstrip("/")
    public = (M.SEARXNG_PUBLIC_URL or "").rstrip("/")
    if not internal or not public or internal == public:
        return payload_json
    return payload_json.replace(internal, public)


def fetch_page(url, max_chars=24000, chunk=1):
    """SSRF-guarded page fetch; the result never exposes the internal URL."""
    return _republicize_internal_url(_fetch_page_impl(url, max_chars, chunk))


def _fetch_page_impl(url, max_chars=24000, chunk=1):
    import ipaddress
    import socket

    from bs4 import BeautifulSoup

    if not url:
        return json.dumps({"url": "", "error": "No URL provided."})
    try:
        chunk = max(1, int(chunk or 1))
    except (TypeError, ValueError):
        chunk = 1
    canon_url = url.split("#", 1)[0] or url
    cached = page_cache.page_get(canon_url)
    if cached is not None:
        print(f"[fetch_page] cache hit for {canon_url}")
        return _doc_result(
            cached["final_url"],
            cached["title"],
            cached["text"],
            max_chars,
            chunk,
            cached["page_images"],
        )
    try:
        parsed = urlparse(url)
        if parsed.scheme not in ("http", "https"):
            return json.dumps(
                {"url": url, "error": "Only http/https URLs are supported."}
            )
        rewritten = _rewrite_search_url_to_internal(parsed)
        if rewritten:
            url = rewritten
        else:
            ip = socket.gethostbyname(parsed.hostname or "")
            addr = ipaddress.ip_address(ip)
            if addr.is_private or addr.is_loopback or addr.is_link_local:
                return json.dumps(
                    {
                        "url": url,
                        "error": "Access to private/internal addresses is not allowed.",
                    }
                )
    except Exception as e:
        return json.dumps({"url": url, "error": f"Invalid URL: {e}"})

    headers = {
        "User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,text/plain;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
    }
    try:
        r = requests.get(url, headers=headers, timeout=15, allow_redirects=True)
        r.raise_for_status()
        raw = getattr(r, "content", None)
        if raw is None:
            raw = r.text.encode("utf-8", errors="replace")
        ctype = r.headers.get("Content-Type", "").lower()
        kind = _detect_doc_type(url, ctype, raw)

        if kind == "pdf":
            text, doc_title, page_images = _parse_pdf(raw, url)
            return _finish_doc(
                canon_url, r.url, doc_title, text, max_chars, chunk,
                doc_type="pdf", page_images=page_images,
            )
        if kind == "csv":
            text, doc_title = _parse_csv(raw, url)
            return _finish_doc(
                canon_url, r.url, doc_title, text, max_chars, chunk, doc_type="csv"
            )
        if kind == "excel":
            text, doc_title = _parse_excel(raw, url)
            return _finish_doc(
                canon_url, r.url, doc_title, text, max_chars, chunk, doc_type="excel"
            )
        if kind == "excel_xls_unsupported":
            return json.dumps(
                {
                    "url": url,
                    "content_type": ctype,
                    "error": "This is a legacy .xls spreadsheet, which is not supported. "
                    "Please retry with a .xlsx or CSV version of the file.",
                }
            )

        if not any(t in ctype for t in _TEXTISH_TYPES):
            return json.dumps(
                {
                    "url": url,
                    "content_type": ctype,
                    "error": "Skipped: page is not readable text content (likely binary/media).",
                }
            )
        if not r.encoding:
            r.encoding = r.apparent_encoding
        soup = BeautifulSoup(r.text, "html.parser")
        for tag in soup(
            [
                "script",
                "style",
                "noscript",
                "svg",
                "nav",
                "footer",
                "header",
                "aside",
                "form",
            ]
        ):
            tag.decompose()
        title = soup.title.get_text(strip=True) if soup.title else ""
        main = soup.find("main") or soup.find("article") or soup.find("body") or soup
        text = main.get_text(separator="\n", strip=True)
        text = "\n".join(line.strip() for line in text.splitlines() if line.strip())
        return _finish_doc(
            canon_url, r.url, title, text, max_chars, chunk, doc_type="web"
        )
    except Exception as e:
        print(f"[fetch_page] Failed: {e}")
        return json.dumps({"url": url, "error": f"Failed to fetch page: {e}"})


def _tool_worker(task_id, sid, tc, image_b64, round_num, tool_index):
    tool_name = tc["function"]["name"]
    try:
        M._dispatch_tool(task_id, sid, tc, image_b64, round_num, tool_index)
    except Exception as e:
        print(f"[tool_worker] Tool '{tool_name}' crashed for task {task_id}: {e}")
        M._event_post(
            "tool_ok",
            task_id,
            tc_id=tc.get("id", ""),
            result=json.dumps({"error": f"Tool {tool_name} failed: {e}"}),
            sid=sid,
            round=round_num,
            tool_index=tool_index,
        )


def _dispatch_tool(task_id, sid, tc, image_b64, round_num, tool_index):
    tool_name = tc["function"]["name"]
    try:
        args = json.loads(tc["function"]["arguments"])
    except Exception:
        args = {}

    with M._data_lock:
        tu = list(M.tasks.get(task_id, {}).get("_tools_used", []))
    has_generated_image = "generate_image" in tu

    if tool_name == "get_user_location":
        if M._client_location:
            result = M._client_location
        else:
            ev = threading.Event()
            M._location_events[task_id] = ev
            M.set_status(task_id, "location_needed")
            ev.wait(timeout=60)
            M._location_events.pop(task_id, None)
            result = (
                M._client_location
                if M._client_location
                else "User denied location access"
            )
        M._event_post(
            "tool_ok",
            task_id,
            tc_id=tc["id"],
            result=result,
            sid=sid,
            round=round_num,
            tool_index=tool_index,
        )
        return

    if tool_name == "read_file":
        file_url = args.get("file_url", "")
        filename = os.path.basename(urlparse(file_url).path)
        fpath = os.path.abspath(os.path.join(M.UPLOADS_DIR, filename))
        if fpath.startswith(os.path.abspath(M.UPLOADS_DIR)) and os.path.exists(fpath):
            text = M.read_file_text(fpath)
            if text:
                markdown_match = re.search(r"\[Markdown saved: (/[^]]+\.md)\]", text)
                if markdown_match:
                    artifact_url = markdown_match.group(1)
                    artifact = {
                        "type": "markdown",
                        "name": os.path.basename(artifact_url),
                        "mime_type": "text/markdown",
                        "url": artifact_url,
                    }
                    with M._data_lock:
                        task = M.tasks.get(task_id)
                        if task:
                            task.setdefault("_artifacts", []).append(artifact)
                result = (
                    f"Content of {file_url}:\n\n{text}\n\n"
                    "This content came from PDF extraction/OCR. Preserve the original "
                    "Unicode text, and use the page headings when quoting or formatting it."
                )
            else:
                result = f"Could not extract text from {file_url}. The file may contain only images."
        else:
            result = f"File not found: {file_url}"
        M._event_post(
            "tool_ok",
            task_id,
            tc_id=tc["id"],
            result=result,
            sid=sid,
            round=round_num,
            tool_index=tool_index,
        )
        return

    if tool_name == "read_image":
        url = args.get("url", "")
        fpath = M.resolve_image_path(url)
        if fpath is None:
            result = json.dumps({"ok": False, "error": f"Image not found: {url}"})
        else:
            result = json.dumps({"ok": True, "image_url": url})
        M._event_post(
            "tool_ok",
            task_id,
            tc_id=tc["id"],
            result=result,
            sid=sid,
            round=round_num,
            tool_index=tool_index,
        )
        return

    if tool_name == "web_search":
        M.set_status(task_id, f"Searching web for: {args.get('query')}...")
        with M._data_lock:
            client_ts = M.tasks.get(task_id, {}).get("_client_timestamp")
        try:
            result = M.web_search(
                args["query"],
                current_time=args.get("current_time"),
                current_location=args.get("current_location"),
            )
        except Exception as e:
            print(f"[web_search] Unhandled exception for task {task_id}: {e}")
            result = json.dumps(
                {"results": [], "query": args.get("query"), "error": str(e)}
            )
        print(f"[web_search] RAW result for task {task_id}: {result[:300]}...")  # DEBUG
        with M._data_lock:
            t = M.tasks.get(task_id)
            if t:
                t.setdefault("_tools_used", []).append(tool_name)
                try:
                    t.setdefault("_search_details", []).append(json.loads(result))
                except Exception:
                    pass
        llm_result = (
            f"Web search results for query '{args.get('query')}'. "
            f"Analyze these search results thoroughly and provide a clear, accurate response based on the findings:\n\n{result}"
        )
        print(
            f"[web_search] LLM-bound result (with analysis instruction) for task {task_id}: {llm_result[:400]}..."
        )  # DEBUG
        M._event_post(
            "tool_ok",
            task_id,
            tc_id=tc["id"],
            result=llm_result,
            sid=sid,
            round=round_num,
            tool_index=tool_index,
        )

    elif tool_name == "fetch_page":
        M.set_status(task_id, f"Fetching page: {args.get('url', '')}...")
        try:
            result = M.fetch_page(args.get("url", ""), chunk=args.get("chunk", 1))
        except Exception as e:
            print(f"[fetch_page] Unhandled exception for task {task_id}: {e}")
            result = json.dumps({"url": args.get("url", ""), "error": str(e)})
        print(f"[fetch_page] Result for task {task_id}: {result[:300]}...")  # DEBUG
        with M._data_lock:
            t = M.tasks.get(task_id)
            if t:
                t.setdefault("_tools_used", []).append(tool_name)
                try:
                    res = json.loads(result)
                    t.setdefault("_search_details", []).append(
                        {
                            "tool": "fetch_page",
                            "url": res.get("url", args.get("url", "")),
                            "title": res.get("title", ""),
                            "content": res.get("content", ""),
                            "error": res.get("error", ""),
                            "retrieved_at": datetime.now().astimezone().isoformat(timespec="seconds"),
                        }
                    )
                except Exception:
                    pass
        llm_result = (
            f"Page content fetched from URL '{args.get('url')}'. "
            f"Use this content to answer the user's question accurately. "
            f"If the content is insufficient or was truncated, you may fetch another page or fall back to the search results:\n\n{result}"
        )
        M._event_post(
            "tool_ok",
            task_id,
            tc_id=tc["id"],
            result=llm_result,
            sid=sid,
            round=round_num,
            tool_index=tool_index,
        )

    elif tool_name == "edit_image":
        M._enqueue_image_job(task_id, sid, tool_name, args, tc, round_num, tool_index)
        return

    elif tool_name == "generate_image":
        if has_generated_image:
            result = json.dumps(
                {"error": "Image generation limit reached for this prompt."}
            )
            M._event_post(
                "tool_ok",
                task_id,
                tc_id=tc["id"],
                result=result,
                sid=sid,
                round=round_num,
                tool_index=tool_index,
            )
        else:
            M._enqueue_image_job(
                task_id, sid, tool_name, args, tc, round_num, tool_index
            )
        return
    elif tool_name == "update_user_context":
        content = args.get("content", "")
        user = ""
        with M._data_lock:
            t = M.tasks.get(task_id)
            if t:
                user = t.get("_user", "")
        if user:
            M.write_user_context(user, content)
            print(f"[context] Updated context for user '{user}' ({len(content)} chars)")
        result = json.dumps({"status": "ok", "saved": bool(user)})
        M._event_post(
            "tool_ok",
            task_id,
            tc_id=tc["id"],
            result=result,
            sid=sid,
            round=round_num,
            tool_index=tool_index,
        )
    elif tool_name == "manage_tasks":
        user = ""
        with M._data_lock:
            t = M.tasks.get(task_id)
            if t:
                user = t.get("_user", "")
        if not user:
            result = json.dumps({"ok": False, "error": "User not found"})
        else:
            result = M.handle_task_tool(user, args)
        M._event_post(
            "tool_ok",
            task_id,
            tc_id=tc["id"],
            result=result,
            sid=sid,
            round=round_num,
            tool_index=tool_index,
        )
    elif tool_name == "track_theme":
        user = ""
        with M._data_lock:
            t = M.tasks.get(task_id)
            if t:
                user = t.get("_user", "")
        if not user:
            result = json.dumps({"ok": False, "error": "User not found"})
        elif user not in M._agent_users:
            result = json.dumps(
                {
                    "ok": False,
                    "error": "track_theme is reserved for the self-chat agent pipeline",
                }
            )
        else:
            result = M.handle_theme_tool(user, args)
        M._event_post(
            "tool_ok",
            task_id,
            tc_id=tc["id"],
            result=result,
            sid=sid,
            round=round_num,
            tool_index=tool_index,
        )
    elif tool_name == "tool_details":
        wanted = [n.strip() for n in str(args.get("name", "")).split(",") if n.strip()]
        known = {t["function"]["name"]: t for t in M.TOOLS_DETAILED}
        with M._data_lock:
            req_user = M.tasks.get(task_id, {}).get("_user", "")
        if req_user not in M._agent_users:
            known = {n: t for n, t in known.items() if n not in M.AGENT_ONLY_TOOLS}
        found = [known[n] for n in wanted if n in known]
        if found:
            result = json.dumps(found)
        else:
            result = json.dumps(
                {
                    "error": "Unknown tool(s)",
                    "requested": wanted,
                    "available": sorted(known),
                }
            )
        M._event_post(
            "tool_ok",
            task_id,
            tc_id=tc["id"],
            result=result,
            sid=sid,
            round=round_num,
            tool_index=tool_index,
        )

    elif mcp_manager.is_mcp_tool(tool_name):
        try:
            result = dispatch_mcp_tool(tool_name, args)
        except Exception as e:
            print(f"[MCP] Tool '{tool_name}' failed: {e}")
            result = json.dumps({"error": f"MCP tool {tool_name} failed: {e}"})

        M._event_post(
            "tool_ok",
            task_id,
            tc_id=tc["id"],
            result=result,
            sid=sid,
            round=round_num,
            tool_index=tool_index,
        )
    else:
        result = json.dumps({"error": f"Unknown tool: {tool_name}"})
        M._event_post(
            "tool_ok",
            task_id,
            tc_id=tc["id"],
            result=result,
            sid=sid,
            round=round_num,
            tool_index=tool_index,
        )
