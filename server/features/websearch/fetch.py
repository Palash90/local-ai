"""Page fetching, SSRF protection, parsing, chunking, and persistence."""

import concurrent.futures
import hashlib
import ipaddress
import json
import os
import socket
from urllib.parse import urlparse, urlunparse

import requests

from server.features.state import M
from server.features.websearch import vector_store as page_cache

def _decode_response_text(raw):
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("latin-1", errors="replace")


def _detect_doc_type(url, content_type, raw):
    """Classify fetched content as ``pdf``/``csv``/``excel`` or ``None``.

    Content-Type is trusted first, then the URL extension, then magic bytes.
    Anything unrecognized returns ``None`` so :func:`fetch_page` declines it as
    binary instead of trying to read it. Legacy ``.xls`` (OLE2 compound
    documents) is detected separately since no parser is available for it.
    """
    ext = os.path.splitext(urlparse(url).path)[1].lower()
    ctype = (content_type or "").split(";")[0].strip().lower()

    if ctype == "application/pdf" or ext == ".pdf" or raw.startswith(b"%PDF-"):
        return "pdf"
    if ctype == "text/csv" or ext == ".csv":
        return "csv"
    # Legacy binary OLE2 compound documents are .xls; we cannot parse those
    # without xlrd, so report them explicitly rather than guessing as binary.
    # Checked before the generic Excel Types because .xls is served as
    # application/vnd.ms-excel just like xlsx templates.
    if ext == ".xls" or raw.startswith(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"):
        return "excel_xls_unsupported"
    if ctype in (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
        "application/excel",
    ) or ext in (".xlsx", ".xlsm", ".xltm", ".xltx"):
        return "excel"
    # Documents served as a generic octet-stream still get parsed if the URL
    # names a known document extension.
    if ctype == "application/octet-stream" and ext in (
        ".pdf",
        ".csv",
        ".xlsx",
        ".xlsm",
        ".xltm",
        ".xltx",
    ):
        return "pdf" if ext == ".pdf" else ("csv" if ext == ".csv" else "excel")
    return None


def _chunk_pages(text, size):
    """Split ``text`` into chunks of up to ``size`` characters."""
    if len(text) <= size:
        return [text]
    return [text[i : i + size] for i in range(0, len(text), size)]


def _doc_result(final_url, title, text, max_chars, chunk=1, page_images=None):
    """Build the ``fetch_page`` JSON payload.

    Long text is split into chunks of up to ``max_chars`` chars and only chunk
    number ``chunk`` is returned; the payload exposes ``total_chunks`` and
    ``next_chunk`` so the agent can page through the rest by calling
    ``fetch_page`` again with ``chunk=2,3,...``. ``page_images`` (rendered PDF
    pages) are attached alongside so a multimodal model can read scanned pages.
    """
    text = (text or "").strip() or "(No readable text content extracted)"
    pages = _chunk_pages(text, max_chars)
    idx = max(0, min(chunk - 1, len(pages) - 1))
    body = pages[idx]
    if page_images:
        body += (
            "\n\n[This PDF has no extractable text — the pages below were rendered "
            "as images. Use the read_image tool on each URL to view a page.]"
        )
    payload = {"url": final_url, "title": title, "content": body}
    if page_images:
        payload["page_images"] = page_images
    if len(pages) > 1:
        payload["chunk"] = idx + 1
        payload["total_chunks"] = len(pages)
        payload["next_chunk"] = idx + 2 if idx + 1 < len(pages) else None
        payload["note"] = (
            f"Page content is split across {len(pages)} chunks. Call fetch_page "
            f"again with chunk={idx + 2} to read the next chunk."
            if payload["next_chunk"]
            else "End of page content."
        )
    return json.dumps(payload, ensure_ascii=False)


def _finish_doc(canon_key, final_url, title, text, max_chars, chunk,
                doc_type="web", page_images=None):
    """Build the fetch payload AND persist the full page under ``canon_key``.

    Only successfully fetched pages reach this point, so the persisted entry is
    guaranteed to have passed the SSRF guard. Pages are stored once with their
    full (pre-chunk) text; every later ``fetch_page(chunk=N)`` re-reads that
    single copy instead of hitting the site again.
    """
    try:
        page_cache.page_put(
            canon_key,
            final_url,
            title or "",
            text or "",
            doc_type=doc_type,
            page_images=page_images,
        )
    except Exception as e:
        print(f"[fetch_page] cache write failed: {e}")
    return _doc_result(final_url, title, text, max_chars, chunk, page_images)


def _render_pdf_pages(doc, url):
    """Render image-only PDF pages (scanned) to PNG files the model can view.

    Rendered files land under ``IMG_PATH/pdf_pages`` and are returned as
    ``/output/pdf_pages/...`` URLs so ``read_image`` / ``resolve_image_path``
    can resolve them. Returns an empty list when nothing could be rendered.
    """

    import hashlib

    slug = hashlib.md5(url.encode("utf-8", errors="replace")).hexdigest()[:12]
    outdir = os.path.join(M.IMG_PATH, "pdf_pages")
    try:
        os.makedirs(outdir, exist_ok=True)
    except OSError as e:
        print(f"[fetch_page] PDF page render dir failed: {e}")
        return []
    urls = []
    for i, page in enumerate(doc):
        if i >= PDF_PAGE_IMAGE_LIMIT:
            break
        try:
            import fitz

            pix = page.get_pixmap(
                matrix=fitz.Matrix(PDF_PAGE_IMAGE_ZOOM, PDF_PAGE_IMAGE_ZOOM)
            )
            fname = f"{slug}-p{i + 1}.png"
            pix.save(os.path.join(outdir, fname))
            urls.append(f"/output/pdf_pages/{fname}")
        except Exception as e:
            print(f"[fetch_page] PDF page {i + 1} render failed: {e}")
            break
    return urls


def _parse_pdf(raw, url):
    """Extract text from a PDF via PyMuPDF.

    Returns ``(text, title, page_images)``. If the PDF has no extractable text
    (scanned pages) the first ``PDF_PAGE_IMAGE_LIMIT`` pages are rendered to PNG
    files and returned as ``/output/pdf_pages/...`` URLs.
    """
    doc_title = os.path.basename(urlparse(url).path) or "PDF document"
    try:
        import fitz

        doc = fitz.open(stream=raw, filetype="pdf")
        parts = []
        total_chars = 0
        for page in doc:
            text = page.get_text("text")
            parts.append(text)
            total_chars += len(text)
            if total_chars > PARSE_PDF_CHARS:
                parts.append("\n...[truncated by size]")
                break
        body = "\n".join(parts).strip()
        page_images = []
        if not body:
            page_images = _render_pdf_pages(doc, url)
            body = "(No extractable text in PDF — the pages are likely scanned images.)"
        doc.close()
        return body, doc_title, page_images
    except Exception as e:
        print(f"[fetch_page] PDF parse failed: {e}")
        return f"(Could not extract text from this PDF: {e})", doc_title, []


def _parse_csv(raw, url):
    """Parse CSV text into pipe-separated rows. Returns ``(text, title)``."""
    import csv
    import io

    doc_title = os.path.basename(urlparse(url).path) or "CSV document"
    try:
        reader = csv.reader(io.StringIO(_decode_response_text(raw)))
        rows = []
        for i, row in enumerate(reader):
            if i >= PARSE_ROW_LIMIT:
                rows.append("...[truncated by row limit]")
                break
            rows.append(" | ".join("" if c is None else c.strip() for c in row))
        body = "\n".join(rows).strip() or "(Empty CSV)"
        return body, doc_title
    except Exception as e:
        print(f"[fetch_page] CSV parse failed: {e}")
        return f"(Could not parse this CSV: {e})", doc_title


def _parse_excel(raw, url):
    """Extract every sheet of an .xlsx workbook as text rows. Returns (text, title)."""
    import io

    from openpyxl import load_workbook

    doc_title = os.path.basename(urlparse(url).path) or "Excel spreadsheet"
    try:
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        blocks = []
        for sheet in wb.worksheets:
            blocks.append(f"### Sheet: {sheet.title}")
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                if row_idx >= PARSE_ROW_LIMIT:
                    blocks.append("...[truncated by row limit]")
                    break
                blocks.append(" | ".join("" if v is None else str(v) for v in row))
        wb.close()
        body = "\n".join(blocks).strip() or "(Empty spreadsheet)"
        return body, doc_title
    except Exception as e:
        print(f"[fetch_page] Excel parse failed: {e}")
        return f"(Could not parse this spreadsheet: {e})", doc_title


def _rewrite_search_url_to_internal(parsed):
    """Map the public SearXNG search URL onto the internal SearXNG instance.

    The public host resolves to a loopback address via /etc/hosts on this
    machine, so fetch_page's SSRF guard would refuse to fetch the very search
    URL that web_search hands out as a citation. Returns None for URLs that
    are not the search endpoint.
    """
    public = urlparse(M.SEARXNG_PUBLIC_URL or "")
    internal = urlparse(M.SEARXNG_URL or "")
    if not (public.netloc and internal.scheme and internal.netloc):
        return None
    base = (public.path or "").rstrip("/")
    path = parsed.path or "/"
    if path != base and not path.startswith(base + "/"):
        return None
    prefix = (internal.path or "").rstrip("/")
    if prefix and path.startswith(prefix):
        path = path[len(prefix):] or "/"
    return urlunparse(internal._replace(path=path, query=parsed.query))


def _republicize_internal_url(payload_json):
    """Replace the internal SearXNG URL with the public one in a tool payload.

    fetch_page rewrites the public search URL onto the internal instance to
    pass the SSRF guard, but the model (and the user) must only ever see the
    public URL — otherwise the private backend address leaks into citations.
    """
    internal = (M.SEARXNG_URL or "").rstrip("/")
    public = (M.SEARXNG_PUBLIC_URL or "").rstrip("/")
    if not internal or not public or internal == public:
        return payload_json
    return payload_json.replace(internal, public)


def fetch_page(url, max_chars=24000, chunk=1):
    """SSRF-guarded page fetch; the result never exposes the internal URL."""
    return _republicize_internal_url(_fetch_page_impl(url, max_chars, chunk))


def _fetch_page_impl(url, max_chars=24000, chunk=1):
    from bs4 import BeautifulSoup

    if not url:
        return json.dumps({"url": "", "error": "No URL provided."})
    try:
        chunk = max(1, int(chunk or 1))
    except (TypeError, ValueError):
        chunk = 1
    canon_url = url.split("#", 1)[0] or url
    cached = page_cache.page_get(canon_url)
    if cached is not None:
        print(f"[fetch_page] cache hit for {canon_url}")
        return _doc_result(
            cached["final_url"],
            cached["title"],
            cached["text"],
            max_chars,
            chunk,
            cached["page_images"],
        )
    try:
        parsed = urlparse(url)
        if parsed.scheme not in ("http", "https"):
            return json.dumps(
                {"url": url, "error": "Only http/https URLs are supported."}
            )
        rewritten = _rewrite_search_url_to_internal(parsed)
        if rewritten:
            url = rewritten
        else:
            ip = socket.gethostbyname(parsed.hostname or "")
            addr = ipaddress.ip_address(ip)
            if addr.is_private or addr.is_loopback or addr.is_link_local:
                return json.dumps(
                    {
                        "url": url,
                        "error": "Access to private/internal addresses is not allowed.",
                    }
                )
    except Exception as e:
        return json.dumps({"url": url, "error": f"Invalid URL: {e}"})

    headers = {
        "User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,text/plain;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
    }
    try:
        r = requests.get(url, headers=headers, timeout=15, allow_redirects=True)
        r.raise_for_status()
        raw = getattr(r, "content", None)
        if raw is None:
            raw = r.text.encode("utf-8", errors="replace")
        ctype = r.headers.get("Content-Type", "").lower()
        kind = _detect_doc_type(url, ctype, raw)

        if kind == "pdf":
            text, doc_title, page_images = _parse_pdf(raw, url)
            return _finish_doc(
                canon_url, r.url, doc_title, text, max_chars, chunk,
                doc_type="pdf", page_images=page_images,
            )
        if kind == "csv":
            text, doc_title = _parse_csv(raw, url)
            return _finish_doc(
                canon_url, r.url, doc_title, text, max_chars, chunk, doc_type="csv"
            )
        if kind == "excel":
            text, doc_title = _parse_excel(raw, url)
            return _finish_doc(
                canon_url, r.url, doc_title, text, max_chars, chunk, doc_type="excel"
            )
        if kind == "excel_xls_unsupported":
            return json.dumps(
                {
                    "url": url,
                    "content_type": ctype,
                    "error": "This is a legacy .xls spreadsheet, which is not supported. "
                    "Please retry with a .xlsx or CSV version of the file.",
                }
            )

        if not any(t in ctype for t in _TEXTISH_TYPES):
            return json.dumps(
                {
                    "url": url,
                    "content_type": ctype,
                    "error": "Skipped: page is not readable text content (likely binary/media).",
                }
            )
        if not r.encoding:
            r.encoding = r.apparent_encoding
        soup = BeautifulSoup(r.text, "html.parser")
        for tag in soup(
            [
                "script",
                "style",
                "noscript",
                "svg",
                "nav",
                "footer",
                "header",
                "aside",
                "form",
            ]
        ):
            tag.decompose()
        title = soup.title.get_text(strip=True) if soup.title else ""
        main = soup.find("main") or soup.find("article") or soup.find("body") or soup
        text = main.get_text(separator="\n", strip=True)
        text = "\n".join(line.strip() for line in text.splitlines() if line.strip())
        return _finish_doc(
            canon_url, r.url, title, text, max_chars, chunk, doc_type="web"
        )
    except Exception as e:
        print(f"[fetch_page] Failed: {e}")
        return json.dumps({"url": url, "error": f"Failed to fetch page: {e}"})


