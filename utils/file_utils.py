def extract_file_text(name, data_b64):
    ext = os.path.splitext(name)[1].lower()
    raw = base64.b64decode(data_b64)
    if ext == ".pdf":
        try:
            import fitz

            doc = fitz.open(stream=raw, filetype="pdf")
            lines = [page.get_text() for page in doc]
            doc.close()
            text = "\n".join(lines)
            return strip_html(text)
        except ImportError:
            with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f:
                f.write(raw)
                tmp = f.name
            try:
                r = subprocess.run(
                    ["pdftotext", tmp, "-"], capture_output=True, text=True, timeout=30
                )
                return r.stdout
            finally:
                os.unlink(tmp)
    elif ext == ".docx":
        from docx import Document

        doc = Document(io.BytesIO(raw))
        return "\n".join(p.text for p in doc.paragraphs)
    elif ext == ".doc":
        with tempfile.NamedTemporaryFile(suffix=".doc", delete=False) as f:
            f.write(raw)
            tmp = f.name
        try:
            r = subprocess.run(
                ["catdoc", tmp], capture_output=True, text=True, timeout=30
            )
            if r.returncode == 0:
                return r.stdout
            r = subprocess.run(
                ["antiword", tmp], capture_output=True, text=True, timeout=30
            )
            return r.stdout
        finally:
            os.unlink(tmp)
    elif ext in (".xls", ".xlsx"):
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        rows = []
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                rows.append("\t".join(str(c) if c is not None else "" for c in row))
        wb.close()
        return "\n".join(rows)
    return ""
